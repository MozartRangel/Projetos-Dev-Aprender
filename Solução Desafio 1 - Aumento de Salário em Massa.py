import openpyxl

#abrir planilha existente
workbook = openpyxl.load_workbook('funcionarios.xlsx')
Sheet_funcionarios = workbook['Employees']

#aumentar o salário em 10% para cada funcionário
for linha in Sheet_funcionarios.iter_rows(min_row=2,values_only=False):
    linha[3].value = linha[3].value * 1.1

for linhas in Sheet_funcionarios.iter_rows(values_only=True):
    print(linhas)

#salvar a planilha atualizada
workbook.save('funcionarios_com_aumento.xlsx')



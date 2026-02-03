from playwright.sync_api import sync_playwright
import openpyxl

workbook = openpyxl.load_workbook('proxies.xlsx')

#abrir o site e coletar os dados depois salvar em uma planilha 
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto('https://proxies-devaprender.netlify.app/')

    #obter todas as linhas
    linhas = page.get_by_role("row").all()

    #ip, port, protocol
    proxies = []
    for linha in linhas[1:]:  # Pular o cabeçalho
        #extrair células
        celulas = linha.get_by_role("cell").all()

        #extrair ip, port, protocol
        ip = celulas[0].inner_text()
        port = celulas[1].inner_text()
        protocol = celulas[4].inner_text()
        
        #adicionar na lista
        proxies.append((ip, port, protocol))

    browser.close()

#salvar na planilha
workbook = openpyxl.Workbook()
#remover a sheet padrão
del workbook['Sheet']
#criar uma nova sheet
planilha = workbook.create_sheet(title='Proxies')
#add cabeçalho
planilha.append(['IP', 'Port', 'Protocol'])
#adicionar os proxies
for proxy in proxies:
    planilha.append(proxy)

workbook.save('proxies.xlsx')
import openpyxl

workbook = openpyxl.load_workbook('esportes.xlsx')

#1. Permitir que o usuário selecione uma página
print(f"páginas disponíveis: {workbook.sheetnames}")
pagina = input("Digite o nome da página que deseja acessar: ").strip()

#2. Permitir que o usuário pesquise por todas as ocorrências que contem um determinado esporte(ex: "Surf") e patrocinador(ex: "Red Bull")
esporte = input("digite o nome do esporte que deseja buscar: ").strip()
patrocinador = input("digite o nome do patrocinador que deseja buscar: ").strip()

for linha in workbook[pagina].iter_rows(min_row=2, values_only=True):
    if linha is not None and linha[2] == esporte and linha[12] == patrocinador:
        print(linha)

